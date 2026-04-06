# -*- coding: utf-8 -*-

import io
import openpyxl
from openpyxl import load_workbook
import os
import unidecode
import pathlib
from datetime import datetime
from odoo.exceptions import ValidationError
from odoo import models, fields, exceptions, api, _
import logging
from io import BytesIO

_logger = logging.getLogger(__name__)

try:
    import csv
except ImportError:
    _logger.debug('Cannot `import csv`.')

try:
    import base64
except ImportError:
    _logger.debug('Cannot `import base64`.')


class gen_journal_entry(models.TransientModel):
    _name = "gen.journal.entry"
    _description = "Generate Journal Entry"

    file_to_upload = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')
    file_name = fields.Char("File Name")

    def find_account_id(self, account_code):
        """
        Find an account by code for the current company, appending zeros if needed.
        """
        if account_code:
            account_ids = self.env['account.account'].search([('code', '=', account_code), ('company_ids', 'in', [self.env.company.id])])
            if account_ids:
                account_id = account_ids[0]
                return account_id
            elif len(account_code) < 8:
                return self.find_account_id(account_code + '0')
        else:
            raise ValidationError(_('Account not found, please verify the current company and accounts') % account_code)

    def find_value_in_line(self, header_fields, line, name, optional=False):
        """
        Find the value in a row based on header name.
        Supports multiple aliases (French or English).
        """
        # Map of accepted column name aliases
        alias_map = {
            'libelle': ['libelle', 'label', 'name', 'description'],
            'compte': ['compte', 'account', 'account_code'],
            'section analytique': ['section analytique', 'analytic', 'analytic_account', 'analytic_account_id'],
            'date': ['date', 'date_maturity', 'journal_date'],
            'debit': ['debit', 'debit_amount'],
            'credit': ['credit', 'credit_amount'],
            'amount_currency': ['amount_currency', 'amount_currency_value', 'amount_devise'],
        }

        # Normalize header names (already lowercase & unaccented earlier)
        header_fields = [h.strip().lower() for h in header_fields]
        possible_names = alias_map.get(name, [name])

        for alias in possible_names:
            if alias in header_fields:
                position = header_fields.index(alias)
                return line[position]

        # If not found
        if not optional:
            raise ValidationError(_("The column '%s' does not exist in your file.") % name)
        return ''

    def check_desc(self, name):
        """Return the name if set, otherwise return '/' as a default description."""
        if name:
            return name
        else:
            return '/'

    def find_account_analytic_id(self, analytic_account_name):
        """Find and return the analytic account ID by name, raise error if not found."""
        analytic_account_id = self.env['account.analytic.account'].search([('name', '=', analytic_account_name)])
        if analytic_account_id:
            analytic_account_id = analytic_account_id[0].id
            return analytic_account_id
        else:
            raise ValidationError(_('Wrong Analytic Account Name %s') % analytic_account_name)

    def find_partner(self, partner_name):
        """Find and return a partner record by name, or None if not found."""
        partner_ids = self.env['res.partner'].search([('name', '=', partner_name)])
        if partner_ids:
            partner_id = partner_ids[0]
            return partner_id
        else:
            partner_id = None

    def check_currency(self, cur_name):
        """Find and return a currency record by name, or None if not found."""
        currency_ids = self.env['res.currency'].search([('name', '=', cur_name)])
        if currency_ids:
            currency_id = currency_ids[0]
            return currency_id
        else:
            currency_id = None
            return currency_id

    def create_import_move_lines(self, values):
        """Validate and map imported move line values to proper record fields.
            - Resolves partner, currency, account, and analytic account IDs.
            - Adjusts debit/credit based on value signs.
            - Ensures name, date, and other fields are correctly formatted.
            - Returns cleaned values ready for account move line creation.
        """
        move_line_obj = self.env['account.move.line']
        move_obj = self.env['account.move']

        if values.get('partner'):
            partner_name = values.get('partner')
            if self.find_partner(partner_name) != None:
                partner_id = self.find_partner(partner_name)
                values.update({'partner_id': partner_id.id})

        if values.get('currency'):
            cur_name = values.get('currency')
            if cur_name != '' and cur_name != None:
                currency_id = self.check_currency(cur_name)
                if currency_id != None:
                    values.update({'currency_id': currency_id.id})
                else:
                    raise ValidationError(_('Currency %s is not  in the system') % cur_name)

        if values.get('name'):
            desc_name = values.get('name')
            name = self.check_desc(desc_name)
            values.update({'name': name})

        if values.get('date_maturity'):
            date = values.get('date_maturity')
            values.update({'date': date})

        if values.get('account_code'):
            account_code = values.get('account_code')
            account_id = self.find_account_id(str(account_code))
            if account_id != None:
                del values['account_code']
                values.update({'account_id': account_id.id})
            else:
                raise ValidationError(_('Wrong Account Code %s') % account_code)

        if values.get('debit') != '' and values.get('debit') != None:
            debit = values.get('debit')
            values.update({'debit': debit})
            if debit < 0:
                values.update({'credit': abs(debit)})
                values.update({'debit': 0.0})
        else:
            values.update({'debit': float('0.0')})

        if values.get('name') == '':
            values.update({'name': '/'})

        if values.get('credit') != '' and values.get('credit') != None:
            values.update({'credit': values.get('credit')})
            if float(values.get('credit')) < 0:
                values.update({'debit': abs(values.get('credit'))})
                values.update({'credit': 0.0})
        else:
            values.update({'credit': float('0.0')})

        if values.get('amount_currency') != '':
            values.update({'amount_currency': values.get('amount_currency')})

        if values.get('analytic_account_id') != '' and values.get('analytic_account_id') != None:
            analytic_account_id = self.find_account_analytic_id(values.get('analytic_account_id'))
            values.update({'analytic_account_id': analytic_account_id})
        else:
            del values['analytic_account_id']
            return values
        return values

    def import_move_lines(self):
        """Import account move lines from CSV or Excel files.

            Validates and parses uploaded files, converts each row into a move line,
            maps necessary fields (partner, account, currency, etc.), and attaches
            them to the selected journal entry. Supports both CSV and XLS/XLSX formats.
        """
        if not self.file_name and not self.file_to_upload:
            raise ValidationError("Please Select File to Upload")

        if self.import_option == 'csv':
            file_extension = pathlib.Path(self.file_name).suffix.lower()
            if file_extension != '.csv':
                raise ValidationError(_("Invalid file extension! Please upload a .csv file."))

            csv_data = base64.b64decode(self.file_to_upload)

            # Detect Excel file accidentally uploaded as CSV
            if csv_data[:2] == b'\x50\x4B' or csv_data[:4] == b'\xD0\xCF\x11\xE0':
                raise ValidationError(_("This file looks like an Excel (.xls/.xlsx). Please choose 'XLS File' option."))

            import chardet
            detected = chardet.detect(csv_data)
            encoding = detected['encoding'] or 'utf-8'

            try:
                data_file = io.StringIO(csv_data.decode(encoding))
            except UnicodeDecodeError:
                raise ValidationError(_("Invalid file format or encoding. Please re-save your file as UTF-8 CSV."))

            data_file.seek(0)
            file_reader = []
            csv_reader = csv.reader(data_file, delimiter=',')

        else:
            if self.import_option == 'xls':
                if pathlib.Path(self.file_name).suffix != '.xls' and pathlib.Path(self.file_name).suffix != '.xlsx':
                    raise ValidationError(_('Invalid file Extention!'))
                decoded_data = base64.b64decode(self.file_to_upload)

                # Verify file really looks like Excel
                if not (decoded_data[:2] == b'\x50\x4B' or decoded_data[:4] == b'\xD0\xCF\x11\xE0'):
                    raise ValidationError(
                        _("Invalid file format! The selected file is not an Excel (.xls/.xlsx). Please choose 'CSV File' instead."))

                wb = openpyxl.load_workbook(filename=BytesIO(decoded_data), read_only=True)
                ws = wb.active
                lines = []
                header_fields = None
                product_obj = self.env['product.product']
                for record in ws.iter_rows(min_row=1, max_row=1, min_col=None, max_col=None, values_only=True):
                    header_fields = [unidecode.unidecode(x.lower()) for x in record if x]
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=True):
                    if self.find_value_in_line(header_fields, record, 'date') != '':
                        values = {
                            'name': self.find_value_in_line(header_fields, record, 'libelle'),
                            'partner_id': False,
                            'analytic_account_id': self.find_value_in_line(header_fields, record, 'section analytique'),
                            'account_code': self.find_value_in_line(header_fields, record, 'compte'),
                            'date_maturity': self.find_value_in_line(header_fields, record, 'date'),
                            'debit': self.find_value_in_line(header_fields, record, 'debit'),
                            'credit': self.find_value_in_line(header_fields, record, 'credit'),
                            'amount_currency': self.find_value_in_line(header_fields, record, 'amount_currency', True),
                            'currency_id': self.env.company.currency_id.id,
                        }
                        if values.get('debit'):
                            values.update({'debit': values.get('debit')})
                        else:
                            values.update({'debit': 0.0})
                        if values.get('credit'):
                            values.update({
                                'credit': values.get('credit')})
                        else:
                            values.update({'credit': 0.0})
                        if (values.get('debit') or values.get('debit') == 0.0) and (
                                values.get('credit') or values.get('credit') == 0.0):
                            values.update({'amount_currency': values.get('debit') - values.get('credit')})
                        if values.get('date_maturity') and type(values.get('date_maturity')) == str:
                            try:
                                date = datetime.strptime(values.get('date_maturity'), '%d/%m/%Y')
                            except:
                                date = datetime.strptime(values.get('date_maturity'), '%d/%m/%y')
                            values.update({'date_maturity': date})
                        else:
                            values.update({'date_maturity': values.get('date_maturity')})
                        account_code = values.get('account_code')
                        if account_code:
                            res = self.create_import_move_lines(values)
                            lines.append((0, 0, res))
            if self._context:
                if self._context.get('active_id'):
                    move_obj = self.env['account.move']
                    move_record = move_obj.browse(self._context.get('active_id'))
                    move_record.write({'line_ids': lines})

        return {'type': 'ir.actions.act_window_close'}
