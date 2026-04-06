# -*- coding: utf-8 -*-

from odoo import models, fields, api
from openpyxl import Workbook
from werkzeug.wrappers import Response
import uuid
import logging
import pdb

excel_headers = {
    'Content-Type': "application/vnd.ms-excel; name='excel'",
    'Content-Disposition': 'inline; filename={}'
}

_logger = logging.getLogger(__name__)


class MrpBom(models.Model):
    _inherit = 'mrp.bom'

    def get_worksheet(self, ws):
        '''This method populates a given worksheet (ws) with data extracted from Bill of Materials (bom_line_ids),
        including product details such as code, quantity, cost, vendor, and product type.
        It also handles sub-BOMs, appending their information to the worksheet as well.'''
        for bom_id in self:
            line_no = 2
            for line in bom_id.bom_line_ids:
                ws['A{}'.format(line_no)] = line.product_id.code
                ws['B{}'.format(line_no)] = line.display_name
                ws['C{}'.format(line_no)] = line.product_qty
                ws['D{}'.format(line_no)] = line.product_tmpl_id.standard_price
                # ws['E{}'.format(line_no)] = line.x_vendor
                ws['F{}'.format(line_no)] = line.product_id.type
                ws['G{}'.format(line_no)] = bom_id.type
                ws['H{}'.format(line_no)] = line.product_id.id
                line_no += 1
                if line.child_bom_id:
                    for sub_line in line.child_bom_id.bom_line_ids:
                        ws['I{}'.format(line_no)] = sub_line.id
                        ws['J{}'.format(line_no)] = sub_line.display_name
                        ws['K{}'.format(line_no)] = sub_line.product_qty
                        # ws['L{}'.format(line_no)] = sub_line.x_vendor
                        ws['M{}'.format(line_no)] = sub_line.product_tmpl_id.standard_price
                        line_no += 1
        return ws

    def mrp_report_cost_review(self):
        '''This method, mrp_report_cost_review, generates an Excel report reviewing the cost of products based on the data retrieved through the get_worksheet method,
        and then serves this report as a response.'''
        wb = Workbook()
        ws = wb.active

        ws['A1'] = "Reference"
        ws['B1'] = "Name"
        ws['C1'] = "Quantity"
        ws['D1'] = "Cost"
        ws['E1'] = "Vendor"
        ws['F1'] = "Product Type"
        ws['G1'] = "BOM Type"
        ws['H1'] = "ID"
        ws['I1'] = "SUB LINE ID"
        ws['J1'] = "Name"
        ws['K1'] = "Quantity"
        ws['L1'] = "Vendor"
        ws['M1'] = "Cost"

        ws = self.get_worksheet(ws)
        fname = "/tmp/{}.xlsx".format(str(uuid.uuid4()).replace("-", ""))
        wb.save(fname)
        fo = open(fname, "rb")
        content = fo.read()
        # Setting proper Content-Disposition header
        excel_headers['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(fname[5:])
        fo.close()  # Close the file
        return Response(content, headers=excel_headers)


class MrpReport(models.TransientModel):
    _name = 'cpe.mrp.report'

    def _get_res_model(self):
        '''This method is responsible for retrieving the name of the active model from the context.
        It returns the active model's name.'''
        return self.env.context.get('active_model', False)

    def _get_res_ids(self):
        '''This method retrieves the IDs of the active records from the context.
        It returns a string containing the active IDs.'''
        active_ids = self.env.context.get('active_ids', [])
        print('active_ids', active_ids)
        return str(active_ids)[1:-1]

    res_model = fields.Char(default=_get_res_model)
    res_ids = fields.Char(default=_get_res_ids)

    def execute_action_mrp_report_cost_review(self):
        '''This method is called when the user executes an action to generate an MRP report cost review.
        It returns an action dictionary, specifying the type of action (ir.actions.act_url) and the URL to be accessed for generating the report.'''
        return {
            "type": "ir.actions.act_url",
            "url": "/cpe/mrp/report/{}?function=mrp_report_cost_review".format(self.id),
            "target": "self",
        }
